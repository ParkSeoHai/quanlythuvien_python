from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.template import loader
from django.urls import reverse
from django.contrib import messages

import uuid
import json
from datetime import datetime
import openpyxl

from library_manager.dtos.CategoryDto import CategoryDto
# Import user dto
from library_manager.dtos.UserDto import UserDto
from library_manager.dtos.AdminDto import AdminDto
from library_manager.dtos.PhieunhapDto import PhieunhapDto
from library_manager.dtos.BookDto import BookDto
from library_manager.dtos.DocgiaDto import DocgiaDto
from library_manager.dtos.ThethuvienDto import ThethuvienDto
from library_manager.dtos.KiemkeDto import KiemkeDto
from library_manager.dtos.CTKiemkeDto import CTKiemkeDto

from library_manager.models import Books, Users, AuthUser, Docgias, Phieunhaps,Phieumuons

from library_manager.dtos.PhieumuonDto import PhieumuonDto
from library_manager.dtos.PhieuhuyDto import PhieuhuyDto

from library_manager.models import Books, Users


# Default view for login page
def index(request):
    template = loader.get_template('login.html')
    return HttpResponse(template.render(request=request))

# Login post request
def loginPost(request):
    if request.method == 'POST':
        # Get email and password from request
        email = request.POST.get('email')
        password = request.POST.get('password')
        # Create user dto
        user = UserDto(email=email, password=password)
        # Response from login function
        response = user.login()

        # If login success
        if response.status is True:
            # Set user id session
            request.session['id_user'] = response.data
            messages.success(request, response.message)
            # Redirect to home page
            return HttpResponseRedirect(reverse('home'))
        else:
            messages.info(request, response.message)
            # Redirect to login page
            return HttpResponseRedirect(reverse('main'))
        
def register(request):
    template = loader.get_template('register.html')
    return HttpResponse(template.render(request=request))

def registerPost(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        repass = request.POST.get('rppassword')
        if UserDto.check_email(email):
            messages.error(request, "Email already registered")
            return HttpResponseRedirect(reverse('register'))
        else:
            if password == repass:
                if len(password) < 6:
                    messages.error(request, "Password must be at least 6 characters long")
                    return HttpResponseRedirect(reverse('register'))
                else:
                    user = UserDto(email=email, password=password)
                    response = user.register()
                    print(response.message)
                    messages.success(request, response.message)
                    return HttpResponseRedirect(reverse('register'))
            else:
                messages.error(request, "Passwords do not match")
                return HttpResponseRedirect(reverse('register'))

# Get user from session
def get_user(request):
    # Get user id session
    id_user = request.session.get('id_user')
    # If user is not logged in, redirect to login page
    if id_user is None:
        return HttpResponseRedirect(reverse('main'))
    else:
        # Get user by id
        response = AdminDto.get_user_by_id(id_user)
        return response.data if response.data else None

# View for home page
def home(request):

    users_response = AdminDto.get_users()
    books_response = UserDto.get_books()
    docgias_response = UserDto.get_docgias()
    phieumuons_response = UserDto.get_phieumuons_all()
    latest_books_response = UserDto.get_books()
    users = users_response.data if users_response.status else []
    books = books_response.data if books_response.status else []
    docgias = docgias_response.data.order_by('-ngay_tao')[:10] if docgias_response.status else []
    phieumuons = phieumuons_response.data if phieumuons_response.status else []
    latest_books = latest_books_response.data.order_by('-ngay_tao')[:10] if latest_books_response.status else []
    context = {
        'user': get_user(request),
        'users': users,
        'books': books,
        'docgias': docgias,
        'phieumuons': phieumuons,
        'latest_books': latest_books,
    }
    template = loader.get_template('home.html')
    return HttpResponse(template.render(context, request))

# View for quanlynguoidung page
def quanlynguoidung(request, tab):
    # Get user id session when user login
    user = get_user(request)
    # Context for template
    context = {
        'user': user,
        'tab': tab,
    }
    # If tab is nguoi-dung
    if tab == 'nguoi-dung':
        # Get all users
        response = AdminDto.get_users()
        context['users'] = response.data
    elif tab == 'doc-gia':
        # Get all thethuviens
        response = AdminDto.get_thethuviens()
        # Context for template
        context['thethuviens'] = response.data
    else:
        print('Tab not found')
        # Message show in template
        messages.info(request, 'Tab not found')

    # Load quanlynguoidung page
    template = loader.get_template('quanlynguoidung/index.html')
    return HttpResponse(template.render(context, request))





    # Load quanlynguoidung page
    template = loader.get_template('quanlynguoidung/index.html')
    return HttpResponse(template.render(context, request))

# Add user view
def addUser(request):
    # Load template
    template = loader.get_template('quanlynguoidung/add.html')
    return HttpResponse(template.render({
        'user': get_user(request)
    }, request))

# Add user post request
def addUserPost(request):
    if request.method == 'POST':
        # Get value
        id = str(uuid.uuid4())  # Generate random id
        name = request.POST.get('name')
        email = request.POST.get('email')
        password = '123456'     # Default password
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        role = int(request.POST.get('role'))
        gender = int(request.POST.get('gender'))
        birthday = request.POST.get('birthday')

        # Create user dto
        user = UserDto(id_user=id, name=name, email=email, password=password, role=role,
                        gender=gender, birthday=birthday, phone_number=phone, address=address)
        print(user.__dict__)

        # Response from add_user function
        response = AdminDto.add_user(user)
        print(response.message)
        if response.status is True:
            # Message
            messages.success(request, response.message)
            return HttpResponseRedirect(reverse('quanlynguoidung', args=('nguoi-dung',)))
        else:
            # Message
            messages.error(request, response.message)
            return HttpResponseRedirect(reverse('addUser'))

# Update user view
def updateUser(request, id):
    # Get user by id
    response = AdminDto.get_user_by_id(id)
    if response.status is False:
        messages.error(request, response.message)

    # Load update user page
    template = loader.get_template('quanlynguoidung/update.html')
    return HttpResponse(template.render({
        'user': get_user(request),
        'user_update': response.data
    }, request))

# Update user post request
def updateUserPost(request):
    if request.method == 'POST':
        # Get value
        id = request.POST.get('id')
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        address = request.POST.get('address')
        role = int(request.POST.get('role'))
        gender = int(request.POST.get('gender'))
        birthday = request.POST.get('birthday')

        # Create user dto
        user = UserDto(id_user=id, name=name, email=email, role=role, gender=gender,
                        birthday=birthday, phone_number=phone, address=address)
        
        print(user.__dict__)

        # Response from update_user function
        response = AdminDto.update_user(user)
        print(response.message)
        if response.status is True:
            messages.success(request, response.message)
            return HttpResponseRedirect(reverse('quanlynguoidung', args=('nguoi-dung',)))
        else:
            messages.error(request, response.message)
            return HttpResponseRedirect(reverse('updateUser', args=(id,)))

# Delete user
def deleteUser(request, id):
    # Response from delete_user function
    response = AdminDto.delete_user(id)
    print(response.message)
    if response.status is False:
        messages.error(request, response.message)
    else:
        messages.success(request, response.message)
    
    return HttpResponseRedirect(reverse('quanlynguoidung', args=('nguoi-dung',)))

# Search user
def searchUser(request, searchInput):
    # Get user
    user = get_user(request)
    # Response from search_user function
    response = AdminDto.search_user(searchInput)
    # Message
    messages.info(request, response.message)
    # Load quanlynguoidung page
    template = loader.get_template('quanlynguoidung/index.html')
    return HttpResponse(template.render({
        'user': user,
        'users': response.data,
        'tab': 'nguoi-dung'
    }, request))

# Add docgia view
def addDocgia(request):
    # Load template
    template = loader.get_template('quanlynguoidung/addDocgia.html')
    return HttpResponse(template.render({
        'user': get_user(request)
    }, request))

# Add docgia post
def addDocgiaPost(request):
    if request.method == 'POST':
        id = str(uuid.uuid4())
        name = request.POST.get('name')
        email = request.POST.get('email')
        address = request.POST.get('address')
        gender = request.POST.get('gender')
        type = request.POST.get('type')
        phone = request.POST.get('phone')
        birthday = request.POST.get('birthday')
        # Get ngay tao is date current
        dateNow = datetime.now()
        ngay_tao = dateNow.strftime("%Y") + "-" + dateNow.strftime("%m") + "-" + dateNow.strftime("%d")

        # Create docgia dto object
        docgiaDto = DocgiaDto(id_docgia=id, name=name, email=email, address=address,
                              gender=int(gender), phone_number=phone, birthday=birthday, ngay_tao=ngay_tao)
        
        # Create thethuvien dto object
        ngay_het_han = str((int(dateNow.strftime("%Y")) + 4)) + "-" + dateNow.strftime("%m") + "-" + dateNow.strftime("%d")
        id_the = str(dateNow.strftime("%Y")) + str(dateNow.strftime("%M")) + str(dateNow.strftime("%S"))
        thethuvienDto = ThethuvienDto(id_the=id_the, type=int(type), ngay_tao=ngay_tao,
                                      ngay_het_han=ngay_het_han)
        
        # Add docgia to database
        response = AdminDto.add_docgia(docgiaDto, thethuvienDto)
        print(response.message)
        if response.status is True:
            # Message show in template
            messages.success(request, response.message)
            return HttpResponseRedirect(reverse('quanlynguoidung', args=('doc-gia',)))
        else:
            # Message show in template
            messages.error(request, response.message)
            return HttpResponseRedirect(reverse('addDocgia'))

# Update doc-gia view
def updateDocgia(request, id):
    # Get thethuvien by id_the
    response = AdminDto.get_thethuvien_by_id(id)
    if response.status is False:
        messages.error(request, response.message)

    # Load update user page
    template = loader.get_template('quanlynguoidung/updateDocgia.html')
    return HttpResponse(template.render({
        'user': get_user(request),
        'ttv': response.data
    }, request))

# Update doc-gia post
def updateDocgiaPost(request):
    if request.method == 'POST':
        id_the = request.POST.get("id-the")
        name = request.POST.get("name")
        email = request.POST.get("email")
        address = request.POST.get("address")
        gender = request.POST.get("gender")
        phone = request.POST.get("phone")
        birthday = request.POST.get("birthday")
        ngay_tao_the = request.POST.get("ngay-tao-the")
        ngay_het_han_the = request.POST.get("ngay-het-han-the")
        type = request.POST.get("type")
        ghi_chu = request.POST.get("ghi-chu")

        # Create docgia dto object
        docgiaDto = DocgiaDto(name=name, email=email, address=address, gender=int(gender),
                              phone_number=phone, birthday=birthday)
        
        # Create thethuvien dto object
        ttvDto = ThethuvienDto(id_the=id_the, ngay_tao=ngay_tao_the, ngay_het_han=ngay_het_han_the,
                               type=type, ghi_chu=ghi_chu)
        
        # Update to database
        response = AdminDto.update_docgia(docgiaDto, ttvDto)
        print(response.message)
        if response.status is True:
            # Message
            messages.success(request, response.message)
            return HttpResponseRedirect(reverse('quanlynguoidung', args=('doc-gia',)))
        else:
            messages.error(request, response.message)
            return HttpResponseRedirect(reverse('updateDocgia', args=(ttvDto.id_the,)))

# Delete doc-gia
def deleteDocgia(request, id):
    # Response from delete_docgia function
    response = AdminDto.delete_docgia(id)
    print(response.message)

    if response.status is True:
        messages.success(request, response.message)
    else:
        messages.error(request, response.message)
    
    return HttpResponseRedirect(reverse('quanlynguoidung', args=('doc-gia',)))

# Search doc-gia
def searchDocgia(request, searchInput):
    # Response from search_docgia function
    response = AdminDto.search_docgia(searchInput)
    # Message
    messages.info(request, response.message)
    # Load quanlynguoidung page
    template = loader.get_template('quanlynguoidung/index.html')
    return HttpResponse(template.render({
        'user': get_user(request),
        'thethuviens': response.data,
        'tab': 'doc-gia'
    }, request))

# Info docgia
def infoDocgia(request, id_the):
    print(id_the)
    # Get info thethuvien by id
    ttvResponse = AdminDto.get_thethuvien_by_id(id_the)
    if ttvResponse.status is True:
        # Get info phieumuons by id_the
        phieumuonResponse = UserDto.get_phieumuonsByThethuvien(ttvResponse.data)

        # Load template
        template = loader.get_template('quanlynguoidung/infoDocgia.html')
        return HttpResponse(template.render({
            'user': get_user(request),
            'ttv': ttvResponse.data,
            'phieumuons': phieumuonResponse.data
        }, request))
    else:
        messages.error(request, ttvResponse.message)
        return HttpResponseRedirect(reverse('quanlynguoidung', args=('doc-gia',)))

# hien thi view quan ly sach
def quanlysach(request):
    # Get user
    user = get_user(request)
    response = UserDto.get_books()
    context = {
        'user': user,
        'books': response.data,
    }
    # Load quanlysach page
    template = loader.get_template('quanlysach/index.html')
    return HttpResponse(template.render(
        context, request))

#xoá sách
def deleteBook(request, id_sach):
    response = UserDto.delete_book(id_sach)
    if response.status is True:
        print(response.message)
        return HttpResponseRedirect(reverse('quanlysach'))
    else:
        print(response.message)
        return HttpResponseRedirect(reverse('quanlysach'))

#them sach
def addBook(request):
    # Get user
    user = get_user(request)
    response = AdminDto.get_categories()
    # Load template
    template = loader.get_template('quanlysach/add.html')
    return HttpResponse(template.render({
        'user': user,
        'categories': response.data,
    }, request))

# Add sach post request
def addBookPost(request):
    if request.method == 'POST':
        # Get value
        id_sach = str(uuid.uuid4())  # Generate random id
        name = request.POST.get('name')
        price = request.POST.get('price')
        quantity = request.POST.get('quantity')
        image = request.POST.get('image')
        author = request.POST.get('author')
        id_category = request.POST.get('category')
        dateNow = datetime.now()
        ngay_tao = dateNow.strftime("%Y") + "-" + dateNow.strftime("%m") + "-" + dateNow.strftime("%d")
        # Create book dto
        book = BookDto(id_sach=id_sach, name=name, price=price, quantity=quantity, image=image,
                        author=author, id_category=id_category, ngay_tao=ngay_tao)
        print(book.__dict__)

        # Response from add_book function
        response = UserDto.add_book(book)
        if response.status is True:
            print(response.message)
            return HttpResponseRedirect(reverse('quanlysach'))
        else:
            print(response.message)
            return HttpResponseRedirect(reverse('addBook'))
        
#update sach
def updateBook(request, id_sach):
    # Get user
    user = get_user(request)
    # Get user by id
    response = AdminDto.get_categories()
    book_update = UserDto.get_bookById(id_sach).data
    # Load update user page
    template = loader.get_template('quanlysach/update.html')
    return HttpResponse(template.render({
        'user': user,
        'book_update': book_update,
        'categories': response.data,
    }, request))

# Update book post request
def updateBookPost(request):
    if request.method == 'POST':
        # Get value
        id_sach = request.POST.get('id_sach')
        name = request.POST.get('name')
        price = request.POST.get('price')
        quantity = request.POST.get('quantity')
        image = request.POST.get('image')
        author = request.POST.get('author')
        id_category = request.POST.get('category')
        # Create user dto
        book = BookDto(id_sach = id_sach,name=name, price=price, quantity=quantity, image=image,
                        author=author, id_category=id_category)
        print(book.__dict__)

        # Response from update_user function
        response = UserDto.update_book(book)
        if response.status is True:
            print(response.message)
            return HttpResponseRedirect(reverse('quanlysach'))
        else:
            print(response.message)
            return HttpResponseRedirect(reverse('updateBook', args=(id_sach,)))

#tim kiem sach
def searchBook(request, searchInput):
    # Get book
    user = get_user(request)
    # Response from search_book function
    response = UserDto.search_book(searchInput)
    # Load quanlysach page
    template = loader.get_template('quanlysach/index.html')
    return HttpResponse(template.render({
        'user': user,
        'books': response.data
    }, request))

# Get book by id from javascript fetch request
def getBook(request):
    if request.method == 'GET':
        id = request.GET.get('id')
        responseDto = UserDto.get_bookById(id[0:len(id)-1])
        # Book response
        book = {
            'id_sach': responseDto.data.id_sach,
            'name': responseDto.data.name,
            'price': responseDto.data.price,
            'quantity': responseDto.data.quantity,
            'image': responseDto.data.image,
            'author': responseDto.data.author,
            'is_delete': responseDto.data.is_delete,
            'id_category': responseDto.data.id_category_id,
            'nameCategory': responseDto.data.id_category.name
        }
        # Response to client
        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': book
        }
        
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Quan ly muon tra sach
def quanlymuontra(request):
    # Load quanlymuontra page
    template = loader.get_template('quanlymuontra/index.html')

    response = UserDto.get_phieumuons()
    return HttpResponse(template.render({
        'user': get_user(request),
        'phieumuons': response.data
    }, request))

# Get books from javascript fetch request - phieumuon
def getBooks_phieumuon(request):
    if request.method == 'GET':
        name = request.GET.get('name')
        response = UserDto.search_booksByName(name[0:len(name)-1])

        books = []
        for item in response.data:
            book = {
                'id_sach': item.id_sach,
                'name': item.name,
                'price': item.price,
                'quantity': item.quantity,
                'image': item.image,
                'author': item.author,
                'is_delete': item.is_delete,
                'id_category': item.id_category.id_category
            }
            books.append(book)

        # Return response to javascript json serializer
        return HttpResponse(json.dumps(books), content_type='application/json')

# Get thethuvien by id_the from javascrip request -phieumuon
def get_info_thethuvienById(request):
    if request.method == 'GET':
        id_the = request.GET.get('id')
        # Response from userdto
        responseDto = AdminDto.get_info_docgiaById_the(id_the[0:len(id_the)-1])
        # Convert to response object
        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': responseDto.data
        }
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Add phieumuon post
def addPhieumuonPost(request):
    if request.method == 'POST':
        data = request.POST.get('data')
        data = json.loads(data)

        bookName = data['bookName']
        id_the = data['id_the']
        ghichu = data['ghichu']
        so_luong = data['so_luong']
        id_user = get_user(request).id_user

        date = datetime.now()
        ngay_tao = f"{date.strftime('%Y')}-{date.strftime('%m')}-{date.strftime('%d')}"
        ngay_hen_tra = f"{date.strftime('%Y')}-{date.strftime('%m')}-{int(date.strftime('%d')) + 7}"

        # Create phieumuon dto
        phieumuonDto = PhieumuonDto(id_phieumuon=str(uuid.uuid4()), ngay_tao=ngay_tao, ngay_hen_tra=ngay_hen_tra,
                                    ghi_chu=ghichu, so_luong=so_luong, id_user=id_user, id_the=id_the)
        
        responseDto = UserDto.add_phieumuon(phieumuonDto, bookName)

        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': responseDto.data
        }
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Get info phieumuon by id_phieumuon]
def get_info_phieumuonById(request, id):
    responseDto = UserDto.get_phieumuonById(id)
    # Convert to response object
    response = {
        'status': responseDto.status,
        'message': responseDto.message,
        'data': responseDto.data
    }
    # Return response to javascript json serializer
    return HttpResponse(json.dumps(response), content_type='application/json')

# Tra sach phieumuon
def trasach(request):
    if request.method == 'POST':
        # Get data
        data = request.POST.get('data')
        data = json.loads(data)

        id_phieumuon = data['id_phieumuon']
        ghi_chu = data['ghi_chu']
        # Get ngay_tra
        date = datetime.now()
        ngay_tra = f"{date.strftime('%Y')}-{date.strftime('%m')}-{date.strftime('%d')}"

        responseDto = UserDto.thuhoi_phieumuon(id_phieumuon, ngay_tra, ghi_chu)
        # Convert to response object
        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': responseDto.data
        }
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Update phieumuon post
def updatePhieumuonPost(request):
    if request.method == 'POST':
        data = request.POST.get('data')
        data = json.loads(data)
        # Get value from data
        id_phieumuon = data['id_phieumuon']
        bookName = data['bookName']
        so_luong = data['so_luong']
        ngay_tao = data['ngay_tao']
        ngay_hen_tra = data['ngay_hen_tra']
        ghi_chu = data['ghi_chu']

        # Create phieumuon dto object
        phieumuonDto = PhieumuonDto(id_phieumuon=id_phieumuon, so_luong=so_luong, ngay_tao=ngay_tao,
                                    ngay_hen_tra=ngay_hen_tra, ghi_chu=ghi_chu)
        # Response from dto
        responseDto = UserDto.update_phieumuon(phieumuonDto, bookName)
        print(responseDto.message)
        # Convert to response object
        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': responseDto.data
        }
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Search phieumuon by id_the
def searchPhieumuonByIdThe(request, id_the):
    responseDto = UserDto.search_phieumuonById_the(id_the)
    # Response object
    response = {
        'status': responseDto.status,
        'message': responseDto.message,
        'data': responseDto.data
    }
    # Return response to javascript json serializer
    return HttpResponse(json.dumps(response), content_type='application/json')

# Delete phieumuon
def deletePhieumuon(request, id):
    response = UserDto.delete_phieumuon(id)
    print(response.message)
    if response.status is True:
        messages.success(request, response.message)
    else:
        messages.error(request, response.message)
    # Redirect to page quanlymuontra
    return HttpResponseRedirect(reverse(quanlymuontra))

def quanlytinhhinhmuontra(request):
    # Get user
    user = get_user(request)
    phieumuons = UserDto.check_phieumuon()
    if phieumuons.data:
        for phieumuon in phieumuons.data:
            ngay_hen_tra = phieumuon.ngay_hen_tra
            today = datetime.now().strftime('%Y-%m-%d')

            if ngay_hen_tra and today:
                a = datetime.strptime(ngay_hen_tra, "%Y-%m-%d")
                b = datetime.strptime(today, "%Y-%m-%d")
                date_muon = int((a - b).days)
        # Load quanlytinhhinhmuontra page
        template = loader.get_template('quanlytinhhinhmuontra/index.html')
        return HttpResponse(template.render({
            'user': user,
            'phieumuons': phieumuons.data,
            'date': date_muon,
        }, request))
    else:
        template = loader.get_template('quanlytinhhinhmuontra/index.html')
        return HttpResponse(template.render({
            'user': user,
            'phieumuons': phieumuons.data,
        }, request))

def quanlytinhhinhDaTra(request):
    # Get user
    user = get_user(request)
    phieumuons = UserDto.check_phieumuonDaTra()
    template = loader.get_template('quanlytinhhinhmuontra/PhieuMuonDaTra.html')
    return HttpResponse(template.render({
        'user': user,
        'phieumuons': phieumuons.data,
    }, request))

# Quan ly kho sach
def quanlykhosach(request):
    # Get user
    user = get_user(request)
    # Load quanlykhosach page
    template = loader.get_template('quanlykhosach/index.html')
    return HttpResponse(template.render({
        'user': user
    }, request))

# Quan ly kho sach - nhap sach
def nhapsach(request):
    # Get user 
    user = get_user(request)
    # Get all phieu nhap
    phieunhapRes = UserDto.get_phieunhaps()
    # Context for template
    context = {
        'user': user,
        'phieunhaps': phieunhapRes.data
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/nhapsach/index.html')
    return HttpResponse(template.render(context, request))

# Lap phieu nhap sach view
def addPhieunhap(request):
    # Get user
    user = get_user(request)
    # Get category
    categoryRes = AdminDto.get_categories()
    # Get books
    bookRes = UserDto.get_books_all()
    # Context for template
    context = {
        'user': user,
        'categories': categoryRes.data,
        'books': bookRes.data
    }
    # Load addPhieunhap page
    template = loader.get_template('quanlykhosach/nhapsach/add.html')
    return HttpResponse(template.render(context, request))

# Cap nhat phieu nhap sach view
def updatePhieunhap(request, id):
    # Get user
    user = get_user(request)
    # Get phieu nhap by id
    pnRes = UserDto.get_phieunhapById(id)
    # Get ctphieunhap by id phieu nhap
    ctphieunhapRes = UserDto.get_ctphieunhapBy_PnhapId(id)
    # List books
    listBooks = []
    for ctpn in ctphieunhapRes.data:
        book = {
            'id': ctpn.id_sach.id_sach,
            'name': ctpn.id_sach.name,
            'price': ctpn.gia_nhap,
            'quantity': ctpn.so_luong,
            'author': ctpn.id_sach.author,
            'category': {
                'id': ctpn.id_sach.id_category_id,
                'name': ctpn.id_sach.id_category.name
            }
        }
        listBooks.append(book)
    # Get categogies
    categoryRes = AdminDto.get_categories()
    # Get books
    bookRes = UserDto.get_books()
    # Context for template
    context = {
        'user': user,
        'phieunhap': pnRes.data,
        'categories': categoryRes.data,
        'books': bookRes.data,
        'listBooks': listBooks
    }
    # Load updatePhieunhap page
    template = loader.get_template('quanlykhosach/nhapsach/update.html')
    return HttpResponse(template.render(context, request))

# Nhap sach post request
def nhapsachPost(request):
    if request.method == 'POST':
        # Get value
        data = request.POST.get('data')
        data = json.loads(data)

        # Create phieu nhap dto
        phieunhap = PhieunhapDto(
            id_phieunhap=str(uuid.uuid4()),
            donvi_cungcap=data['dvcc'],
            ngay_nhap=data['ngaynhap'],
            ly_do_nhap=data['lydo'],
            id_user=request.session.get('id_user')
        )
        
        # Create list book dto
        books = []
        for book in data['books']:
            book = BookDto(
                id_sach=str(uuid.uuid4()),
                name=book['name'],
                price=float(book['price']),
                quantity=int(book['quantity']),
                image='https://plus.unsplash.com/premium_photo-1667251760532-85310936c89a?q=80&w=1887&auto=format&fit=crop&ixlib=rb-4.0.3&ixid=M3wxMjA3fDB8MHxwaG90by1wYWdlfHx8fGVufDB8fHx8fA%3D%3D',
                author=book['author'],
                id_category=book['category']['id'],
                ngay_tao=data['ngaynhap']
            )
            books.append(book)
        
        # Response from nhap_sach function
        responseDto = UserDto.nhap_sach(phieunhap, books)
        print(responseDto.message)

        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': responseDto.data
        }
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Update phieu nhap
def updatePhieunhapPost(request):
    if request.method == 'POST':
        # Get value
        data = request.POST.get('data')
        data = json.loads(data)

        print(data)

        # Create phieu nhap dto
        pn = PhieunhapDto(
            id_phieunhap=data['id_phieunhap'],
            donvi_cungcap=data['dvcc'],
            ngay_nhap=data['ngaynhap'],
            ly_do_nhap=data['lydo'],
            id_user=request.session.get('id_user')
        )
        
        # Create list book dto
        books = []
        for book in data['books']:
            book = BookDto(
                id_sach=book['id'],
                name=book['name'],
                price=float(book['price']),
                quantity=int(book['quantity']),
                image='https://plus.unsplash.com/premium_photo-1667251760532-85310936c89a?q=80&w=1887&auto=format&fit=crop&ixlib=rb-4.0.3&ixid=M3wxMjA3fDB8MHxwaG90by1wYWdlfHx8fGVufDB8fHx8fA%3D%3D',
                author=book['author'],
                id_category=book['category']['id'],
                ngay_tao=data['ngaynhap']
            )
            books.append(book)

        # Get book deletes
        book_deletes = data['book_deletes']

        # Response from nhap_sach function
        responseDto = UserDto.update_phieunhap(pn, books, book_deletes)
        print(responseDto.message)

        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': responseDto.data
        }

        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Delete phieu nhap
def deletePhieunhap(request, id):
    # Response from delete_phieunhap function
    response = UserDto.delete_phieunhap(id)
    print(response.message)
    if response.status is True:
        messages.success(request, response.message)
    else:
        messages.error(request, response.message)

    return HttpResponseRedirect(reverse('nhapsachIndex'))

# Search phieu nhap by date
def searchPhieunhapByDate(request, dateFrom, dateTo):
    # Get user
    user = get_user(request)
    # Response from search_phieunhap_by_date function
    response = UserDto.search_phieunhap_by_date(dateFrom, dateTo)
    # Context for template
    context = {
        'user': user,
        'phieunhaps': response.data,
        'dateFrom': dateFrom,
        'dateTo': dateTo
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/nhapsach/index.html')
    return HttpResponse(template.render(context, request))

# Info phieu nhap sach view
def infoPhieunhap(request, id):
    # Get user
    user = get_user(request)
    # Get phieu nhap by id
    phieunhapRes = UserDto.get_phieunhapById(id)
    # Get ctphieunhap by id phieu nhap
    ctphieunhapRes = UserDto.get_ctphieunhapBy_PnhapId(id)
    # Context for template
    context = {
        'user': user,
        'phieunhap': phieunhapRes.data,
        'ctphieunhaps': ctphieunhapRes.data
    }
    # Load infoPhieunhap page
    template = loader.get_template('quanlykhosach/nhapsach/info.html')
    return HttpResponse(template.render(context, request))

# Quan ly kho sach - huy sach
def huysach(request):
    # Get list phieuhuys
    phieuhuyResponse = UserDto.get_phieuhuys()

    context = {
        'user': get_user(request),
        'phieuhuys': phieuhuyResponse.data
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/huysach/index.html')
    return HttpResponse(template.render(context, request))

# Quan ly kho sach - huy sach - add view
def addPhieuhuy(request):
    # Get all books
    books = UserDto.get_books_all().data
    context = {
        'user': get_user(request),
        'books': books
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/huysach/add.html')
    return HttpResponse(template.render(context, request))

# Quan ly kho sach - huy sach - update view
def updatePhieuhuy(request, id):
    # Get all books
    books = UserDto.get_books_all().data
    # Get phieuhuy
    phieuhuyRes = UserDto.get_phieuhuyById(id)
    # Get ctphieuhuys
    ctphieuhuyRes = UserDto.get_ctphieuhuyBy_PhuyId(id)
    # List books
    listBooks = []
    for ctph in ctphieuhuyRes.data:
        book = {
            'id_ctph': ctph.id_ctphieuhuy,
            'id_sach': ctph.id_sach.id_sach,
            'name': ctph.id_sach.name,
            'category': ctph.id_sach.id_category.name,
            'author': ctph.id_sach.author,
            'quantity': ctph.so_luong,
            'quantityCurr': ctph.id_sach.quantity,
            'note': ctph.ly_do_huy,
        }
        listBooks.append(book)

    context = {
        'user': get_user(request),
        'phieuhuy': phieuhuyRes.data,
        'books': books,
        'listBooks': listBooks
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/huysach/update.html')
    return HttpResponse(template.render(context, request))

# Add phieuhuy post
def huysachPost(request):
    if request.method == 'POST':
        # Get value
        data = request.POST.get('data')
        data = json.loads(data)
        # Get ngayhuy by date now
        date = datetime.now()
        ngay_huy = f'{date.strftime('%Y')}-{date.strftime('%m')}-{date.strftime('%d')}'
        # Create phieu huy dto
        phieuhuyDto = PhieuhuyDto(
            id_phieuhuy=str(uuid.uuid4()),
            ngay_huy=ngay_huy,
            id_user=request.session.get('id_user'),
        )
        # Get ct phieuhuy
        ctphieuhuys = []
        for ct in data['books']:
            ctObject = {
                'id_ctphieuhuy': str(uuid.uuid4()),
                'id_sach': ct['id_sach'],
                'so_luong': int(ct['quantity']),
                'ly_do_huy': ct['note']
            }
            ctphieuhuys.append(ctObject)

        responseDto = UserDto.huy_sach(phieuhuyDto, ctphieuhuys)
        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': responseDto.data
        }
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Update phieuhuy post
def updatePhieuhuyPost(request):
    if request.method == 'POST':
        data = request.POST.get('data')
        data = json.loads(data)
        # Create phieuhuy object dto
        phieuhuyDto = PhieuhuyDto(
            id_phieuhuy=data['id_phieuhuy'],
            ngay_huy=data['ngay_huy']
        )
        # Get ct phieuhuy
        ctphieuhuys = []
        for ct in data['books']:
            ctObject = {
                'id_ctphieuhuy': str(uuid.uuid4()),
                'id_sach': ct['id_sach'],
                'so_luong': int(ct['quantity']),
                'ly_do_huy': ct['note']
            }
            ctphieuhuys.append(ctObject)
        # Get ctph delete
        ctph_deletes = data['ctph_deletes']

        # Response from userdto
        responseDto = UserDto.update_phieuhuy(phieuhuyDto, ctphieuhuys, ctph_deletes)
        # Convert to response object
        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': responseDto.data
        }
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Delete phieuhuy
def deletePhieuhuy(request, id):
    response = UserDto.delete_phieuhuy(id)
    print(response.message)
    if response.status is True:
        messages.success(request, response.message)
    else:
        messages.error(request, response.message)
    # Redirect to page quan-ly-kho-sach / huy-sach
    return HttpResponseRedirect(reverse('huysachIndex'))

# Info phieuhuy view
def infoPhieuhuy(request, id):
    # Get phieuhuy by id
    phieuhuy = UserDto.get_phieuhuyById(id).data
    # Get all ctphieuhuys
    ctphieuhuys = UserDto.get_ctphieuhuyBy_PhuyId(phieuhuy.id_phieuhuy).data
    context = {
        'user': get_user(request),
        'phieuhuy': phieuhuy,
        'ctphieuhuys': ctphieuhuys
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/huysach/info.html')
    return HttpResponse(template.render(context, request))

# Search phieuhuy by date
def searchPhieuhuyByDate(request, dateFrom, dateTo):
    # Response from search_phieuhuyby_date function
    response = UserDto.search_phieuhuy_by_date(dateFrom, dateTo)
    # Context for template
    context = {
        'user': get_user(request),
        'phieuhuys': response.data,
        'dateFrom': dateFrom,
        'dateTo': dateTo
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/huysach/index.html')
    return HttpResponse(template.render(context, request))

# Quan ly kho sach - kiem ke
def kiemkeView(request):
    # Get all kiemke
    kiemkeResponse = UserDto.get_phieuKiemkes()
    context = {
        'user': get_user(request),
        'kiemkes': kiemkeResponse.data
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/kiemke/index.html')
    return HttpResponse(template.render(context, request))

# Add kiem ke view
def addKiemkeView(request):
    context = {
        'user': get_user(request),
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/kiemke/add.html')
    return HttpResponse(template.render(context, request))

# Get data file excel kiemke from request js
def get_dataKiemkeFile(request):
    if request.method == 'POST':
        data = request.POST.get('data')
        data = json.loads(data)
        fileName = data['fileName']
        file_path = r'D:\IT3-EAUT\Python\btl\{}'.format(fileName)

        # Set url file
        workbook = openpyxl.load_workbook(filename=file_path)
        workbook.sheetnames
        sheet = workbook.active
        # Get list book
        list_book = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None and row[1] is None:
                break
            book = {
                'id_sach': row[0],
                'name': row[1],
                'category': row[2],
                'author': row[3],
                'so_luong_bandau': row[4],
                'so_luong_kiemke': row[5],
                'chenh_lech': int(row[5]) - int(row[4])
            }
            list_book.append(book)
        
        # Response to client
        response = {
            'status': False,
            'data': list_book
        }

        if len(list_book) > 0:
            response['status'] = True
            response['message'] = f'Get data from file {fileName} success'
        else:
            response['message'] = f'Get data from file {fileName} failed'
            
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Add phieu kiemke post
def add_phieuKiemkePost(request):
    if request.method == 'POST':
        data = request.POST.get('data')
        data = json.loads(data)
        # Get ngay_tao
        date = datetime.now()
        ngay_tao = f'{date.strftime('%Y')}-{date.strftime('%m')}-{date.strftime('%d')}'
        # Create kiemke dto object
        kiemkeDto = KiemkeDto(
            id_kiemke=str(uuid.uuid4()),
            ngay_tao=ngay_tao,
            ly_do=data['ly_do'],
            file_kiemke=data['file_kiemke'],
            id_user=request.session.get('id_user'),
        )
        # Get list ct_kiemke dto
        ct_kiemkes = []
        for ct in data['ct_kiemkes']:
            ct_kiemke = CTKiemkeDto(
                id_ctkiemke=str(uuid.uuid4()),
                id_sach=ct['id_sach'],
                so_luong_bandau=int(ct['so_luong_bandau']),
                so_luong_kiemke=int(ct['so_luong_kiemke'])
            )
            ct_kiemkes.append(ct_kiemke)

        # Response from user dto kiemke
        responseDto = UserDto.kiemke(kiemkeDto, ct_kiemkes)
        response = {
            'status': responseDto.status,
            'message': responseDto.message,
            'data': responseDto.data
        }
        # Return response to javascript json serializer
        return HttpResponse(json.dumps(response), content_type='application/json')

# Info kiemke view
def info_kiemkeView(request, id):
    # Get phieu kiemke
    kiemkeResponse = UserDto.get_phieuKiemkeById(id)
    # Kiemke not found
    if kiemkeResponse.status is False:
        messages.error(request, kiemkeResponse.message)
        return HttpResponseRedirect(reverse('kiemkeView'))
    
    # Get ct_kiemke
    ct_kiemkeResponse = UserDto.get_ctKiemkesById_kiemke(id)
    # Get context
    context = {
        'user': get_user(request),
        'kiemke': kiemkeResponse.data,
        'ct_kiemkes': ct_kiemkeResponse.data
    }
    # Load nhapsach page
    template = loader.get_template('quanlykhosach/kiemke/info.html')
    return HttpResponse(template.render(context, request))

# Logout
def logout(request):
    # Delete user id session
    if 'id_user' in request.session:
        del request.session['id_user']
    # Redirect to login page
    return HttpResponseRedirect(reverse('main'))

def quanlydanhmuc(request):
    # Get user
    user = get_user(request)
    response = AdminDto.get_categories()
    # Load quanlydanhmuc page
    template = loader.get_template('quanlydanhmuc/index.html')
    return HttpResponse(template.render({
        'user': user,
        'categories' :response.data
    }, request))

def addUsertoCategory(request):
    # Get user
    user = get_user(request)
    # Load template
    template = loader.get_template('quanlydanhmuc/add.html')
    return HttpResponse(template.render({
        'user': user
    }, request))

def addCategoryPost(request):
    if request.method == 'POST':
        # Get value
        id = str(uuid.uuid4())  # Generate random id
        name = request.POST.get('name')
        description = request.POST.get('description')
        isDelete = 0   # Default type for delete

        # Create Category dto
        category = CategoryDto(id_category=id, name=name, description=description, is_delete=isDelete)
        print(category.__dict__)

        # Response from add_Category function
        response = AdminDto.add_category(category)
        print(response.message)
        if response.status is True:
            messages.success(request, response.message)
            return HttpResponseRedirect(reverse('quanlydanhmuc'))
        else:
            messages.error(request, response.message)
            return HttpResponseRedirect(reverse('addUsertoCategory'))
        
def deleteCategory(request, id):
    # Response from delete_user function
    response = AdminDto.delete_category(id)
    print(response.message)
    if response.status is True:
        messages.success(request, response.message)
        return HttpResponseRedirect(reverse('quanlydanhmuc'))
    else:
        messages.error(request, response.message)
        return HttpResponseRedirect(reverse('quanlydanhmuc'))
    
def updateCategory(request, id):
    # Get user
    user = get_user(request)
    # Get user by id
    category_update = AdminDto.get_category_by_id(id).data
    # Load update user page
    template = loader.get_template('quanlydanhmuc/update.html')
    return HttpResponse(template.render({
        'user': user,
        'category_update': category_update
    }, request))

def updateCategoryPost(request):
    if request.method == 'POST':
        # Get value
        id = request.POST.get('id')
        name = request.POST.get('name')
        desciption = request.POST.get('description')
        # Update user dto
        category = CategoryDto(id_category=id, name=name,description=desciption, is_delete=0)
        print(category.__dict__)
        response = AdminDto.update_category(category)
        print(response.message)
        # Response from update_user function
        if response.status is True:
            messages.success(request, response.message)
            return HttpResponseRedirect(reverse('quanlydanhmuc'))
        else:
            messages.error(request, response.message)
            return HttpResponseRedirect(reverse('updateCategory', args=(id,)))

def searchCategories(request, searchInput):
    # Get user
    user = get_user(request)
    # Response from search_user function
    response = AdminDto.search_category(searchInput)
    # Load quanlynguoidung page
    template = loader.get_template('quanlydanhmuc/index.html')
    return HttpResponse(template.render({
        'user': user,
        'categories': response.data
    }, request))

def searchThongKeSachTK(request, searchInput):
    user = get_user(request)
    response = UserDto.searchSachTK(searchInput)
    template = loader.get_template('quanlykhosach/sachtonkho/index.html')
    return HttpResponse(template.render({
        'user': user,
        'thongkesachs': response.data
    }, request))

# Thong ke
def ThongKeNhapHuy(request):
    user = get_user(request)
    response = UserDto.ThongKePhieuNhap()
    response_huy = UserDto.ThongKePhieuHuy()
    template = loader.get_template('thongke/thongkenhaphuy.html')
    return HttpResponse(template.render({
        'user': user,
        'Nhap': response.data,
        'Huy': response_huy.data,
    }))

def ThongKeTongKho(request):
    user = get_user(request)
    response = UserDto.thongkesach()
    template = loader.get_template('thongke/thongketonkho.html')
    return HttpResponse(template.render({
        'user': user,
        'thongkesachs': response.data
    }, request))